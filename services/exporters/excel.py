from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill


def create_excel(search_data, schools):
    """엑셀 파일 생성"""
    wb = Workbook()
    ws = wb.active
    ws.title = "검색결과"

    # 스타일 정의
    title_font = Font(bold=True, size=14)
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font_white = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # 제목
    ws.merge_cells('A1:F1')
    ws['A1'] = "울산교육청 재난안전 학교검색 결과"
    ws['A1'].font = title_font
    ws['A1'].alignment = Alignment(horizontal='center')

    # 검색 정보
    ws['A3'] = f"검색일시: {search_data['datetime']}"
    ws['A4'] = f"검색주소: {search_data['address']}"
    ws['A5'] = f"검색좌표: 위도 {search_data['lat']}, 경도 {search_data['lng']}"
    ws['A6'] = f"검색반경: {search_data['radius']}km"
    ws['A7'] = f"검색결과: {len(schools)}개교"

    # 헤더
    headers = ['순번', '학교명', '학교급', '학급수', '학생수', '거리(km)', '도로명주소', '지번주소']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=9, column=col, value=header)
        cell.font = header_font_white
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')

    # 데이터
    for idx, school in enumerate(schools, 1):
        row_num = 9 + idx
        ws.cell(row=row_num, column=1, value=idx).border = thin_border
        ws.cell(row=row_num, column=1).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=2, value=school['name']).border = thin_border
        ws.cell(row=row_num, column=3, value=school['type']).border = thin_border
        ws.cell(row=row_num, column=3).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=4, value=school.get('classes', 0)).border = thin_border
        ws.cell(row=row_num, column=4).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=5, value=school.get('students', 0)).border = thin_border
        ws.cell(row=row_num, column=5).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=6, value=school['distance']).border = thin_border
        ws.cell(row=row_num, column=6).alignment = Alignment(horizontal='center')
        ws.cell(row=row_num, column=7, value=school.get('road_address', '')).border = thin_border
        ws.cell(row=row_num, column=8, value=school.get('address', '')).border = thin_border

    # 열 너비 조정
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 45
    ws.column_dimensions['H'].width = 45

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer
